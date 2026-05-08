# -*- coding: utf-8 -*-
"""
자금일보 검증기 추가 샘플 생성 스크립트 (양식 다양성 확장).

웹 조사로 파악한 한국 기업의 일반적 자금일보 양식 2종을 추가:

  · 양식 A — 자금일계표(시계열 Ledger형)
      yesform "자금일계표(공통서식)" 패턴 — 거래마다 한 행, 시간순 ledger
      컬럼: 일자 | 시간 | 구분 | 계정과목 | 거래처 | 적요 | 수입 | 지출
      → 자금일보_샘플3.xlsx, .pdf / 은행거래내역_샘플3.xlsx, .pdf

  · 양식 B — 항목별 분류형(대분류 + 결제수단 추가)
      bizforms / yesform "일일 자금일보" 패턴 — 대분류·계정과목·거래처별로 분류
      컬럼: 대분류 | 계정과목 | 거래처 | 적요 | 결제수단 | 입금 | 출금 | 비고
      → 자금일보_샘플4.xlsx, .pdf / 은행거래내역_샘플4.xlsx, .pdf

검증 시나리오:
  · 샘플3 (양식 A, 2025-04-15): 9건 모두 1:1 매칭. 입금 75M / 출금 46M
  · 샘플4 (양식 B, 2025-05-08): 1:1 7건 + N:1 1건(일반경비 5M = 1.5+1.5+2)
                                입금 50.5M / 출금 55M

각 샘플은 simulate_validator(엑셀) / simulate_pdf_path(PDF) 시뮬레이터로
검증 가능하도록 검증기 파서 제약(헤더 키워드, 합계 행 필터, 거래 분류
키워드)을 만족하게 구성됨.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

OUT_DIR = Path(__file__).resolve().parent

THIN = Side(style="thin", color="888888")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="E8EEF5")
SECTION_FILL = PatternFill("solid", fgColor="F4F6F9")
INCOME_FILL = PatternFill("solid", fgColor="ECFDF5")
EXPENSE_FILL = PatternFill("solid", fgColor="FEF2F2")


def register_korean_font():
    candidates = [
        ("MalgunGothic", r"C:\Windows\Fonts\malgun.ttf"),
        ("MalgunGothicBold", r"C:\Windows\Fonts\malgunbd.ttf"),
        ("NanumGothic", r"C:\Windows\Fonts\NanumGothic.ttf"),
    ]
    registered = []
    for name, path in candidates:
        if Path(path).exists():
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                registered.append(name)
            except Exception as e:
                print(f"  [WARN] font register failed: {name}: {e}")
    if not registered:
        raise RuntimeError("한글 폰트를 찾지 못했습니다 (Malgun/Nanum)")
    return registered[0], (registered[1] if len(registered) > 1 else registered[0])


KOR_FONT, KOR_FONT_BOLD = register_korean_font()


def styled(font=None, size=10, bold=False, align="LEFT"):
    return ParagraphStyle(
        "k",
        fontName=KOR_FONT_BOLD if bold else (font or KOR_FONT),
        fontSize=size,
        leading=size * 1.4,
        alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}[align],
    )


# ════════════════════════════════════════════════════════════════════
# 시나리오 데이터
# ════════════════════════════════════════════════════════════════════

# ── 양식 A (샘플3): 자금일계표 시계열 Ledger ──
# 자금일보와 은행이 1:1로 정확히 일치
SCENARIO_A_DATE = "2025-04-15"
SCENARIO_A_OPENING = 80_000_000
SCENARIO_A_LEDGER = [
    # (시간, 구분, 계정과목, 거래처, 적요, 수입, 지출)
    ("09:00:14", "수입", "매출수금",     "(주)한솔테크",   "1분기 매출 회수",      28_000_000, 0),
    ("09:30:42", "수입", "매출수금",     "(주)대한건설",   "4월 매출채권 회수",    22_000_000, 0),
    ("10:15:08", "수입", "외상대 회수",  "(주)미래상사",   "매출채권 회수",        25_000_000, 0),
    ("11:30:33", "지출", "외상대 결제",  "(주)동방산업",   "원자재 매입 결제",     0, 12_000_000),
    ("13:20:50", "지출", "외상대 결제",  "(주)성광물산",   "부품 매입 결제",       0, 8_000_000),
    ("14:00:11", "지출", "인건비",       "임직원 일괄",    "4월분 급여 지급",      0, 18_000_000),
    ("15:30:25", "지출", "일반경비",     "한빛빌딩",       "임대료 납부",          0, 5_000_000),
    ("16:00:18", "지출", "일반경비",     "KT",             "통신비 납부",          0, 1_500_000),
    ("16:30:47", "지출", "일반경비",     "오피스월드",     "사무용품 구입",        0, 1_500_000),
]

# ── 양식 B (샘플4): 분류형 + N:1 합산 매칭 ──
SCENARIO_B_DATE = "2025-05-08"
SCENARIO_B_OPENING = 200_000_000

# 자금일보 8건 (일반경비를 묶어서 5M, 은행에서는 3건으로 분해 → N:1)
SCENARIO_B_FUND = [
    # (대분류, 계정과목, 거래처, 적요, 결제수단, 입금, 출금)
    ("수입", "매출수금",      "(주)A상사",      "매출채권 회수",     "입금이체",   30_000_000, 0),
    ("수입", "매출수금",      "(주)B산업",      "매출채권 회수",     "입금이체",   20_000_000, 0),
    ("수입", "예금이자",      "KB은행",          "정기예금 이자 입금", "자동입금",   500_000,    0),
    ("지출", "외상대 결제",   "(주)C기업",      "원자재 매입 결제",  "이체출금",   0,           8_000_000),
    ("지출", "외상대 결제",   "(주)D기업",      "부품 매입 결제",    "이체출금",   0,          12_000_000),
    ("지출", "인건비",         "임직원 일괄",    "5월 급여 지급",     "자동이체",   0,          25_000_000),
    ("지출", "일반경비",       "사무관리비",     "사무용품·통신·임대료", "복합",     0,           5_000_000),
    ("지출", "세금",           "강남세무서",     "부가세 납부",        "자동이체",   0,           5_000_000),
]

# 은행 10건 (일반경비 5M이 1.5+1.5+2로 분해)
SCENARIO_B_BANK = [
    # (시간, 입금, 출금, 적요, 거래처, 결제수단)
    ("09:15:20", 30_000_000, 0,           "매출수금",      "(주)A상사",       "입금이체"),
    ("09:45:33", 20_000_000, 0,           "매출수금",      "(주)B산업",       "입금이체"),
    ("10:00:11", 500_000,    0,           "예금이자 입금",  "KB은행",          "자동입금"),
    ("11:00:45", 0,          8_000_000,   "외상대 결제",   "(주)C기업",       "이체출금"),
    ("11:01:22", 0,          12_000_000,  "외상대 결제",   "(주)D기업",       "이체출금"),
    ("13:30:08", 0,          25_000_000,  "5월 급여 지급",  "임직원 일괄",     "자동이체"),
    ("14:10:55", 0,          1_500_000,   "사무용품 구입",  "오피스월드",      "카드결제"),
    ("14:25:14", 0,          1_500_000,   "통신비 자동이체", "KT",              "자동이체"),
    ("15:00:39", 0,          2_000_000,   "임대료 납부",    "한빛빌딩",        "이체출금"),
    ("16:30:02", 0,          5_000_000,   "부가세 납부",    "강남세무서",      "자동이체"),
]


# ════════════════════════════════════════════════════════════════════
# 1) 양식 A — Excel
# ════════════════════════════════════════════════════════════════════
def build_form_a_fund_excel():
    """샘플3 자금일보(엑셀) — 자금일계표 시계열 Ledger.

    검증기 parseFundExcel 호환:
      · 헤더 키워드: '구분', '계정과목', '수입', '지출', '잔액'
      · 합계 행은 '합계' 키워드로 자동 필터됨
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "자금일계표"

    # 메타 (상단 박스 — 헤더 매칭 점수 낮게 유지)
    ws["A1"] = "일일 자금일계표"
    ws["A1"].font = Font(size=15, bold=True)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")

    meta = [
        ("회사명", "(주)동행기업", "작성일자", SCENARIO_A_DATE),
        ("기초시재(전일이월)", SCENARIO_A_OPENING,
         "기말시재(당일잔액)", SCENARIO_A_OPENING + 75_000_000 - 46_000_000),
    ]
    for r_off, row in enumerate(meta, start=2):
        for c_off, val in enumerate(row, start=1):
            cell = ws.cell(row=r_off, column=c_off, value=val)
            if c_off in (1, 3):
                cell.font = Font(bold=True)
                cell.fill = SECTION_FILL
            elif isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"

    # 표 헤더
    header_row = 5
    headers = ["일자", "시간", "구분", "계정과목", "거래처", "적요", "수입", "지출"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")

    # 거래 행
    for i, (t, kind, acct, vendor, memo, dep, wd) in enumerate(
        SCENARIO_A_LEDGER, start=header_row + 1
    ):
        vals = [SCENARIO_A_DATE, t, kind, acct, vendor, memo, dep, wd]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col in (7, 8) and isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0;-#,##0;\"-\""
            elif col in (1, 2, 3):
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
            if kind == "수입":
                if col == 3:
                    cell.fill = INCOME_FILL
            elif kind == "지출":
                if col == 3:
                    cell.fill = EXPENSE_FILL

    # 합계 행
    tot_row = header_row + 1 + len(SCENARIO_A_LEDGER)
    sum_dep = sum(r[5] for r in SCENARIO_A_LEDGER)
    sum_wd = sum(r[6] for r in SCENARIO_A_LEDGER)
    ws.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=6)
    ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=tot_row, column=7, value=sum_dep).number_format = "#,##0"
    ws.cell(row=tot_row, column=8, value=sum_wd).number_format = "#,##0"
    for col in range(1, 9):
        c = ws.cell(row=tot_row, column=col)
        c.font = Font(bold=True)
        c.fill = SECTION_FILL
        c.border = BOX
        if col in (7, 8):
            c.alignment = Alignment(horizontal="right")

    # 열 너비
    widths = {"A": 13, "B": 11, "C": 8, "D": 16, "E": 18, "F": 22, "G": 14, "H": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = OUT_DIR / "자금일보_샘플3.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")
    return out


def build_form_a_bank_excel():
    """샘플3 은행거래내역(엑셀) — 자금일보와 1:1 일치하는 9건."""
    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"

    ws["A1"] = "거래내역 조회"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"], ws["B2"] = "계좌번호", "456-789012-03-001"
    ws["A3"], ws["B3"] = "예금주", "(주)동행기업"
    ws["A4"], ws["B4"] = "조회기간", f"{SCENARIO_A_DATE} ~ {SCENARIO_A_DATE}"

    headers = ["거래일시", "입금액", "출금액", "잔액", "적요", "거래처", "은행"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")

    bal = SCENARIO_A_OPENING
    for i, (t, kind, acct, vendor, memo, dep, wd) in enumerate(
        SCENARIO_A_LEDGER, start=7
    ):
        bal = bal + dep - wd
        date_str = f"{SCENARIO_A_DATE} {t}"
        vals = [date_str, dep, wd, bal, memo, vendor, "신한은행"]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col in (2, 3, 4) and isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

    tot_row = 7 + len(SCENARIO_A_LEDGER)
    ws.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=tot_row, column=2, value=sum(r[5] for r in SCENARIO_A_LEDGER)).number_format = "#,##0"
    ws.cell(row=tot_row, column=3, value=sum(r[6] for r in SCENARIO_A_LEDGER)).number_format = "#,##0"
    for col in range(1, 8):
        c = ws.cell(row=tot_row, column=col)
        c.font = Font(bold=True)
        c.fill = SECTION_FILL
        c.border = BOX

    widths = {"A": 22, "B": 14, "C": 14, "D": 16, "E": 22, "F": 22, "G": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = OUT_DIR / "은행거래내역_샘플3.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")
    return out


# ════════════════════════════════════════════════════════════════════
# 2) 양식 A — PDF
# ════════════════════════════════════════════════════════════════════
def build_form_a_fund_pdf():
    """샘플3 자금일보(PDF) — 시계열 Ledger.

    PDF 파서(parseDailyFundPlan)는 라인 단위로 항목명+숫자를 추출.
    각 거래 라인에 한 종류 금액(수입 또는 지출)만 두고 0은 '-'로 표시 →
    extract_valid_amounts가 거래액 하나만 잡도록 유도.
    """
    out = OUT_DIR / "자금일보_샘플3.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=18 * mm, bottomMargin=15 * mm,
    )

    title = styled(size=17, bold=True, align="CENTER")
    sub = styled(size=10, align="CENTER")

    story = [
        Paragraph("일일 자금일계표", title),
        Spacer(1, 4 * mm),
        Paragraph(
            f"회사명: (주)동행기업   |   작성일자: {SCENARIO_A_DATE}   |   단위: 원",
            sub,
        ),
        Spacer(1, 6 * mm),
    ]

    closing = SCENARIO_A_OPENING + 75_000_000 - 46_000_000
    meta = Table(
        [["기초시재(전일이월)", f"{SCENARIO_A_OPENING:,}원",
          "기말시재(당일잔액)", f"{closing:,}원"]],
        colWidths=[45 * mm, 45 * mm, 45 * mm, 45 * mm],
    )
    meta.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), KOR_FONT_BOLD, 11),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#E8EEF5")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#E8EEF5")),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("ALIGN", (3, 0), (3, 0), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(meta)
    story.append(Spacer(1, 6 * mm))

    fmt = lambda n: f"{n:,}" if isinstance(n, int) and n > 0 else "-"
    headers = ["일자", "시간", "구분", "계정과목", "거래처", "적요", "수입", "지출"]
    rows_pdf = [headers] + [
        [SCENARIO_A_DATE, t, kind, acct, vendor, memo, fmt(dep), fmt(wd)]
        for (t, kind, acct, vendor, memo, dep, wd) in SCENARIO_A_LEDGER
    ]
    sum_dep = sum(r[5] for r in SCENARIO_A_LEDGER)
    sum_wd = sum(r[6] for r in SCENARIO_A_LEDGER)
    # 합계는 표 아래 별도 텍스트로 — 표 안에 두면 PDF 라인 단위 파서가
    # "합계 X X" 한 줄을 transaction으로 오인식하므로.

    table_rows = []
    for i, r in enumerate(rows_pdf):
        cells = []
        for j, c in enumerate(r):
            align = "RIGHT" if j in (6, 7) else ("CENTER" if j in (0, 1, 2) else "LEFT")
            bold = (i == 0)
            cells.append(Paragraph(str(c), styled(size=8, bold=bold, align=align)))
        table_rows.append(cells)

    tbl = Table(
        table_rows,
        colWidths=[22 * mm, 18 * mm, 14 * mm, 25 * mm, 30 * mm, 32 * mm, 22 * mm, 22 * mm],
    )
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    # 수입/지출 행 색 구분
    for i, r in enumerate(SCENARIO_A_LEDGER, start=1):
        kind = r[1]
        if kind == "수입":
            style_cmds.append(("BACKGROUND", (2, i), (2, i), colors.HexColor("#ECFDF5")))
        else:
            style_cmds.append(("BACKGROUND", (2, i), (2, i), colors.HexColor("#FEF2F2")))
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        f"※ 본 자금일계표 — 수입 합계: {sum_dep:,}원 / 지출 합계: {sum_wd:,}원",
        styled(size=9),
    ))

    doc.build(story)
    print(f"  [OK] {out.name}")
    return out


def build_form_a_bank_pdf():
    """샘플3 은행거래내역(PDF) — 자금일보와 1:1 매칭되는 9건."""
    out = OUT_DIR / "은행거래내역_샘플3.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    title = styled(size=16, bold=True, align="CENTER")
    sub = styled(size=9, align="CENTER")
    story = [
        Paragraph("거래내역 조회", title),
        Spacer(1, 3 * mm),
        Paragraph(
            f"계좌번호: 456-789012-03-001   |   예금주: (주)동행기업   |   "
            f"조회기간: {SCENARIO_A_DATE} ~ {SCENARIO_A_DATE}",
            sub,
        ),
        Spacer(1, 5 * mm),
    ]

    headers = ["거래일시", "입금액", "출금액", "잔액", "적요", "거래처"]
    bal = SCENARIO_A_OPENING
    data = [headers]
    for (t, kind, acct, vendor, memo, dep, wd) in SCENARIO_A_LEDGER:
        bal = bal + dep - wd
        data.append([
            f"{SCENARIO_A_DATE} {t}",
            f"{dep:,}" if dep else "0",
            f"{wd:,}" if wd else "0",
            f"{bal:,}",
            memo,
            vendor,
        ])
    data.append([
        "합계",
        f"{sum(r[5] for r in SCENARIO_A_LEDGER):,}",
        f"{sum(r[6] for r in SCENARIO_A_LEDGER):,}",
        "", "", "",
    ])

    table_rows = []
    for i, row in enumerate(data):
        cells = []
        for j, c in enumerate(row):
            align = "RIGHT" if j in (1, 2, 3) else ("CENTER" if j == 0 else "LEFT")
            bold = (i == 0 or i == len(data) - 1)
            cells.append(Paragraph(str(c), styled(size=8, bold=bold, align=align)))
        table_rows.append(cells)

    tbl = Table(
        table_rows,
        colWidths=[34 * mm, 23 * mm, 23 * mm, 28 * mm, 35 * mm, 39 * mm],
    )
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F4F6F9")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    doc.build(story)
    print(f"  [OK] {out.name}")
    return out


# ════════════════════════════════════════════════════════════════════
# 3) 양식 B — Excel
# ════════════════════════════════════════════════════════════════════
def build_form_b_fund_excel():
    """샘플4 자금일보(엑셀) — 분류형 + 결제수단 컬럼."""
    wb = Workbook()
    ws = wb.active
    ws.title = "자금일보"

    ws["A1"] = "일일 자금일보"
    ws["A1"].font = Font(size=15, bold=True)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")

    closing = SCENARIO_B_OPENING + 50_500_000 - 55_000_000
    meta = [
        ("회사명", "(주)글로벌테크", "작성일자", SCENARIO_B_DATE),
        ("기초시재", SCENARIO_B_OPENING, "기말시재", closing),
    ]
    for r_off, row in enumerate(meta, start=2):
        for c_off, val in enumerate(row, start=1):
            cell = ws.cell(row=r_off, column=c_off, value=val)
            if c_off in (1, 3):
                cell.font = Font(bold=True)
                cell.fill = SECTION_FILL
            elif isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"

    header_row = 5
    headers = ["대분류", "계정과목", "거래처", "적요", "결제수단", "입금", "출금", "비고"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")

    for i, (cat, acct, vendor, memo, pay, dep, wd) in enumerate(
        SCENARIO_B_FUND, start=header_row + 1
    ):
        vals = [cat, acct, vendor, memo, pay, dep, wd, ""]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col in (6, 7) and isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0;-#,##0;\"-\""
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")
                if cat == "수입":
                    cell.fill = INCOME_FILL
                else:
                    cell.fill = EXPENSE_FILL
            else:
                cell.alignment = Alignment(horizontal="left")

    tot_row = header_row + 1 + len(SCENARIO_B_FUND)
    sum_dep = sum(r[5] for r in SCENARIO_B_FUND)
    sum_wd = sum(r[6] for r in SCENARIO_B_FUND)
    ws.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=5)
    ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=tot_row, column=6, value=sum_dep).number_format = "#,##0"
    ws.cell(row=tot_row, column=7, value=sum_wd).number_format = "#,##0"
    for col in range(1, 9):
        c = ws.cell(row=tot_row, column=col)
        c.font = Font(bold=True)
        c.fill = SECTION_FILL
        c.border = BOX

    widths = {"A": 10, "B": 16, "C": 18, "D": 22, "E": 12, "F": 14, "G": 14, "H": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = OUT_DIR / "자금일보_샘플4.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")
    return out


def build_form_b_bank_excel():
    """샘플4 은행거래내역(엑셀) — 자금일보 8행과 1:1 7건 + N:1 1건(일반경비)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"

    ws["A1"] = "거래내역 조회"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"], ws["B2"] = "계좌번호", "789-012345-04-002"
    ws["A3"], ws["B3"] = "예금주", "(주)글로벌테크"
    ws["A4"], ws["B4"] = "조회기간", f"{SCENARIO_B_DATE} ~ {SCENARIO_B_DATE}"

    headers = ["거래일시", "입금액", "출금액", "잔액", "적요", "거래처", "결제수단"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")

    bal = SCENARIO_B_OPENING
    for i, (t, dep, wd, memo, vendor, pay) in enumerate(SCENARIO_B_BANK, start=7):
        bal = bal + dep - wd
        date_str = f"{SCENARIO_B_DATE} {t}"
        vals = [date_str, dep, wd, bal, memo, vendor, pay]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col in (2, 3, 4) and isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

    tot_row = 7 + len(SCENARIO_B_BANK)
    ws.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=tot_row, column=2, value=sum(r[1] for r in SCENARIO_B_BANK)).number_format = "#,##0"
    ws.cell(row=tot_row, column=3, value=sum(r[2] for r in SCENARIO_B_BANK)).number_format = "#,##0"
    for col in range(1, 8):
        c = ws.cell(row=tot_row, column=col)
        c.font = Font(bold=True)
        c.fill = SECTION_FILL
        c.border = BOX

    widths = {"A": 22, "B": 14, "C": 14, "D": 16, "E": 22, "F": 22, "G": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = OUT_DIR / "은행거래내역_샘플4.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")
    return out


# ════════════════════════════════════════════════════════════════════
# 4) 양식 B — PDF
# ════════════════════════════════════════════════════════════════════
def build_form_b_fund_pdf():
    """샘플4 자금일보(PDF) — 분류형."""
    out = OUT_DIR / "자금일보_샘플4.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=18 * mm, bottomMargin=15 * mm,
    )

    title = styled(size=17, bold=True, align="CENTER")
    sub = styled(size=10, align="CENTER")

    closing = SCENARIO_B_OPENING + 50_500_000 - 55_000_000
    story = [
        Paragraph("일일 자금일보", title),
        Spacer(1, 4 * mm),
        Paragraph(
            f"회사명: (주)글로벌테크   |   작성일자: {SCENARIO_B_DATE}   |   단위: 원",
            sub,
        ),
        Spacer(1, 6 * mm),
    ]

    meta = Table(
        [["기초시재", f"{SCENARIO_B_OPENING:,}원",
          "기말시재", f"{closing:,}원"]],
        colWidths=[45 * mm, 45 * mm, 45 * mm, 45 * mm],
    )
    meta.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), KOR_FONT_BOLD, 11),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#E8EEF5")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#E8EEF5")),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("ALIGN", (3, 0), (3, 0), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(meta)
    story.append(Spacer(1, 6 * mm))

    fmt = lambda n: f"{n:,}" if isinstance(n, int) and n > 0 else "-"
    headers = ["대분류", "계정과목", "거래처", "적요", "결제수단", "입금", "출금"]
    rows_pdf = [headers] + [
        [cat, acct, vendor, memo, pay, fmt(dep), fmt(wd)]
        for (cat, acct, vendor, memo, pay, dep, wd) in SCENARIO_B_FUND
    ]
    sum_dep = sum(r[5] for r in SCENARIO_B_FUND)
    sum_wd = sum(r[6] for r in SCENARIO_B_FUND)
    # 합계 행은 표 안에 넣지 않고 표 아래 별도 텍스트로 (PDF 라인 파서 호환)

    table_rows = []
    for i, r in enumerate(rows_pdf):
        cells = []
        for j, c in enumerate(r):
            align = "RIGHT" if j in (5, 6) else ("CENTER" if j in (0, 4) else "LEFT")
            bold = (i == 0)
            cells.append(Paragraph(str(c), styled(size=8, bold=bold, align=align)))
        table_rows.append(cells)

    tbl = Table(
        table_rows,
        colWidths=[16 * mm, 26 * mm, 30 * mm, 38 * mm, 20 * mm, 24 * mm, 24 * mm],
    )
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for i, r in enumerate(SCENARIO_B_FUND, start=1):
        cat = r[0]
        bg = "#ECFDF5" if cat == "수입" else "#FEF2F2"
        style_cmds.append(("BACKGROUND", (0, i), (0, i), colors.HexColor(bg)))
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        f"※ 본 자금일보 — 수입 합계: {sum_dep:,}원 / 지출 합계: {sum_wd:,}원",
        styled(size=9),
    ))
    story.append(Paragraph(
        f"※ 일반경비(5,000,000)는 사무용품·통신비·임대료 합산 항목으로 "
        f"은행에서는 3건으로 분리되어 있을 수 있음 (N:1 매칭 검증 케이스)",
        styled(size=9),
    ))

    doc.build(story)
    print(f"  [OK] {out.name}")
    return out


def build_form_b_bank_pdf():
    """샘플4 은행거래내역(PDF) — 일반경비 3건 분해."""
    out = OUT_DIR / "은행거래내역_샘플4.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    title = styled(size=16, bold=True, align="CENTER")
    sub = styled(size=9, align="CENTER")
    story = [
        Paragraph("거래내역 조회", title),
        Spacer(1, 3 * mm),
        Paragraph(
            f"계좌번호: 789-012345-04-002   |   예금주: (주)글로벌테크   |   "
            f"조회기간: {SCENARIO_B_DATE} ~ {SCENARIO_B_DATE}",
            sub,
        ),
        Spacer(1, 5 * mm),
    ]

    headers = ["거래일시", "입금액", "출금액", "잔액", "적요", "거래처"]
    bal = SCENARIO_B_OPENING
    data = [headers]
    for (t, dep, wd, memo, vendor, pay) in SCENARIO_B_BANK:
        bal = bal + dep - wd
        data.append([
            f"{SCENARIO_B_DATE} {t}",
            f"{dep:,}" if dep else "0",
            f"{wd:,}" if wd else "0",
            f"{bal:,}",
            memo,
            vendor,
        ])
    data.append([
        "합계",
        f"{sum(r[1] for r in SCENARIO_B_BANK):,}",
        f"{sum(r[2] for r in SCENARIO_B_BANK):,}",
        "", "", "",
    ])

    table_rows = []
    for i, row in enumerate(data):
        cells = []
        for j, c in enumerate(row):
            align = "RIGHT" if j in (1, 2, 3) else ("CENTER" if j == 0 else "LEFT")
            bold = (i == 0 or i == len(data) - 1)
            cells.append(Paragraph(str(c), styled(size=8, bold=bold, align=align)))
        table_rows.append(cells)

    tbl = Table(
        table_rows,
        colWidths=[34 * mm, 23 * mm, 23 * mm, 28 * mm, 35 * mm, 39 * mm],
    )
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F4F6F9")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    doc.build(story)
    print(f"  [OK] {out.name}")
    return out


# ════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("[양식 A — 자금일계표 시계열 Ledger형] 생성 중...")
    build_form_a_fund_excel()
    build_form_a_bank_excel()
    build_form_a_fund_pdf()
    build_form_a_bank_pdf()
    print()
    print("[양식 B — 분류형(대분류/결제수단)] 생성 중...")
    build_form_b_fund_excel()
    build_form_b_bank_excel()
    build_form_b_fund_pdf()
    build_form_b_bank_pdf()
    print()
    print("=" * 60)
    print("생성 완료. test_samples/ 폴더에서 다음 파일을 확인하세요:")
    print("  · 자금일보_샘플3.xlsx, .pdf  +  은행거래내역_샘플3.xlsx, .pdf  (양식 A)")
    print("  · 자금일보_샘플4.xlsx, .pdf  +  은행거래내역_샘플4.xlsx, .pdf  (양식 B)")
