# -*- coding: utf-8 -*-
"""
샘플6 자동 생성기 경로 시뮬레이션 — 4개 은행 → 1개 자금일보.

자금일보 자동 생성기는 은행 거래내역 N개 파일을 누적 업로드하면
generateFundReport()가 1개 자금일보를 생성. 흐름:

  1. 각 은행 파일 → handleGenBankFile()
     · PDF: extractPdfTextLineAware → parseBankPDF → parseBankRows
     · Excel: sheet_to_json → autoDetectMapping → parseBankRows
     → genBankFiles 누적
  2. generateFundReport()
     · 4개 파일의 transactions 모두 합치기
     · 계좌별 그룹화 → 합계/잔액
     · 주요 거래 (상위 20% 등) 추출
     · HTML/Excel 자금일보 출력

이 스크립트는 위 흐름을 Python으로 재현해서:
  · 4개 은행 파일에서 정확히 13건 거래가 누적되는지
  · 합계가 입금 135M / 출금 238M 으로 맞는지
  · 계좌별 그룹화가 정확한지
  · 주요 거래 추출이 작동하는지
"""
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

from simulate_pdf_path import parse_bank_pdf

OUT = Path(__file__).resolve().parent

EXPECTED = {
    "신한": {"count": 5, "deposit": 115_000_000, "withdrawal": 178_500_000},
    "국민": {"count": 3, "deposit": 0, "withdrawal": 52_000_000},
    "하나": {"count": 3, "deposit": 0, "withdrawal": 7_000_000},
    "우리": {"count": 2, "deposit": 20_000_000, "withdrawal": 500_000},
}
TOTAL_DEPOSIT = 135_000_000
TOTAL_WITHDRAWAL = 238_000_000


def extract_pdf_lineaware(path):
    """index2_pdf.html의 extractPdfTextLineAware 시뮬레이션."""
    with pdfplumber.open(path) as pdf:
        return "\n".join(page.extract_text() for page in pdf.pages)


def parse_bank_xlsx(path):
    """index2_pdf.html의 Excel 분기 + autoDetectMapping + parseBankRows 재현."""
    from simulate_validator import parse_bank_excel
    wb = load_workbook(path)
    ws = wb.active
    raw = [list(row) for row in ws.iter_rows(values_only=True)]
    return parse_bank_excel(raw)["transactions"]


def parse_bank_file(path, ext):
    """is_pdf 인지에 따라 다른 파서 호출."""
    if ext == "pdf":
        text = extract_pdf_lineaware(path)
        parsed = parse_bank_pdf(text)
        # parseBankPDF의 rows: [date, dep, wd, bal, desc]
        return [
            {
                "date": r[0], "deposit": r[1], "withdrawal": r[2],
                "balance": r[3], "description": r[4],
            }
            for r in parsed["rows"]
        ]
    else:
        return parse_bank_xlsx(path)


def calc_highlight_threshold(transactions, method="top20"):
    """index2_pdf.html의 calcHighlightThreshold 재현."""
    amounts = sorted(
        [max(t["deposit"], t["withdrawal"]) for t in transactions
         if max(t["deposit"], t["withdrawal"]) > 0]
    )
    if not amounts:
        return 0
    pct = 0.9 if method == "top10" else 0.8
    idx = int(len(amounts) * pct)
    return amounts[min(idx, len(amounts) - 1)]


def run(ext):
    print()
    print("=" * 72)
    print(f"  자동 생성기 경로 시뮬레이션 — {ext.upper()} 파일 4개 → 1개 자금일보")
    print("=" * 72)

    all_transactions = []
    file_summaries = []

    for key in ["신한", "국민", "하나", "우리"]:
        fname = f"은행거래내역_샘플6_{key}.{ext}"
        path = OUT / fname
        txns = parse_bank_file(path, ext)
        for t in txns:
            t["bankLabel"] = fname
        all_transactions.extend(txns)

        dep_sum = sum(t["deposit"] for t in txns)
        wd_sum = sum(t["withdrawal"] for t in txns)
        exp = EXPECTED[key]
        ok = (
            len(txns) == exp["count"]
            and dep_sum == exp["deposit"]
            and wd_sum == exp["withdrawal"]
        )
        mark = "✓" if ok else "✗"
        file_summaries.append((key, len(txns), dep_sum, wd_sum, ok))
        print(f"  {mark} {fname:<40} {len(txns)}건  "
              f"입금 {dep_sum:>13,}  출금 {wd_sum:>13,}")

    # 통합 결과
    total_dep = sum(t["deposit"] for t in all_transactions)
    total_wd = sum(t["withdrawal"] for t in all_transactions)
    print()
    print(f"  [통합] 총 거래 건수: {len(all_transactions)}건 (기대 13건)")
    print(f"  [통합] 입금 합계: {total_dep:>15,}원 (기대 {TOTAL_DEPOSIT:,}원)")
    print(f"  [통합] 출금 합계: {total_wd:>15,}원 (기대 {TOTAL_WITHDRAWAL:,}원)")

    # 그룹화 (파일별)
    print()
    print("  [계좌별 그룹화]")
    groups = {}
    for t in all_transactions:
        key = t["bankLabel"]
        if key not in groups:
            groups[key] = {"deposit": 0, "withdrawal": 0, "count": 0}
        groups[key]["deposit"] += t["deposit"]
        groups[key]["withdrawal"] += t["withdrawal"]
        groups[key]["count"] += 1
    for key, g in groups.items():
        print(f"    · {key:<40} {g['count']}건  "
              f"입금 {g['deposit']:>13,}  출금 {g['withdrawal']:>13,}")

    # 주요 거래 (상위 20%)
    threshold = calc_highlight_threshold(all_transactions, "top20")
    significant = [
        t for t in all_transactions
        if max(t["deposit"], t["withdrawal"]) >= threshold and threshold > 0
    ]
    print()
    print(f"  [주요 거래] 임계금액 {threshold:,}원 이상 — {len(significant)}건")
    for t in significant:
        amt = max(t["deposit"], t["withdrawal"])
        kind = "입금" if t["deposit"] > t["withdrawal"] else "출금"
        print(f"    ⭐ {kind} {amt:>13,}원  [{t['bankLabel']}] {t['description']}")

    # 검증
    all_ok = (
        len(all_transactions) == 13
        and total_dep == TOTAL_DEPOSIT
        and total_wd == TOTAL_WITHDRAWAL
        and all(s[4] for s in file_summaries)
    )
    print()
    if all_ok:
        print(f"  ✓ {ext.upper()} 자동 생성기 경로 — 모든 거래 정상 인식 (13건, 합계 일치)")
    else:
        print(f"  ✗ {ext.upper()} 자동 생성기 경로 — 검증 실패")
    return all_ok


if __name__ == "__main__":
    excel_ok = run("xlsx")
    pdf_ok = run("pdf")
    print()
    print("=" * 72)
    print(f"  최종 결과: Excel {'OK' if excel_ok else 'FAIL'} | PDF {'OK' if pdf_ok else 'FAIL'}")
    print("=" * 72)
