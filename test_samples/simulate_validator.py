# -*- coding: utf-8 -*-
"""
검증기의 JS 파서 로직을 Python으로 시뮬레이션해서 버그를 재현/검증한다.
parseFundExcel, parseBankRows, crossValidate의 핵심 흐름을 그대로 옮긴다.
"""
import re
from openpyxl import load_workbook
from pathlib import Path

OUT = Path(__file__).resolve().parent


def parse_amount(val):
    if val is None or val == "":
        return 0
    s = str(val).strip()
    s = re.sub(r"[,\s원₩]", "", s)
    m = re.match(r"^\((.+)\)$", s)
    if m:
        s = "-" + m.group(1)
    try:
        return abs(float(s))
    except ValueError:
        return 0


def parse_fund_excel(raw):
    """index2_pdf.html의 parseFundExcel 로직을 그대로 재현."""
    result = {"transactions": [], "date": None, "totals": {}}
    if not raw or len(raw) < 2:
        return result

    # ── 헤더 행 탐지: 키워드 매칭 셀 수가 가장 많은 행 선택 ──
    header_kw_re = re.compile(
        r"잔액|잔고|입금|출금|수입|지출|차변|대변|전일|당일|기초|기말|"
        r"구\s*분|항\s*목|계\s*정|계\s*좌|적\s*요|비\s*고"
    )
    header_idx = -1
    best = 0
    for i in range(min(len(raw), 15)):
        cells = [str(c if c is not None else "").strip().lower() for c in raw[i]]
        cnt = 0
        for c in cells:
            if not c or len(c) > 20:
                continue
            if re.match(r"^\d[\d,. ]*$", c):
                continue
            if header_kw_re.search(c):
                cnt += 1
        if cnt > best:
            best = cnt
            header_idx = i
    if header_idx < 0 or best < 2:
        for i in range(min(len(raw), 10)):
            non_empty = sum(1 for c in raw[i] if c is not None and str(c).strip())
            if non_empty >= 4 and i > 0:
                header_idx = i
                break
    if header_idx < 0:
        header_idx = 0

    print(f"[Fund] headerIdx={header_idx}, headers={raw[header_idx]}")
    headers = [str(c if c is not None else "").strip().lower() for c in raw[header_idx]]

    # ── 컬럼 자동 감지 ──
    col_account = -1
    col_bank = -1
    col_prev = -1
    col_deposit = -1
    col_withdrawal = -1
    col_balance = -1
    for i, h in enumerate(headers):
        if "계좌" in h or "계정" in h or "구분" in h or "항목" in h:
            col_account = i
        if "은행" in h or "금융기관" in h:
            col_bank = i
        if "전일" in h or "전잔" in h or "기초" in h:
            col_prev = i
        if ("입금" in h or "수입" in h or "증가" in h) and "출금" not in h:
            col_deposit = i
        if ("출금" in h or "지출" in h or "감소" in h) and "입금" not in h:
            col_withdrawal = i
        if (("당일" in h and "잔액" in h) or "당잔" in h or "기말" in h
                or ("잔액" in h and "전일" not in h and "전잔" not in h)):
            col_balance = i

    if col_balance < 0:
        for i, h in enumerate(headers):
            if ("잔액" in h or "잔고" in h) and i != col_prev:
                col_balance = i
                break

    if col_deposit < 0 and col_withdrawal < 0:
        numeric_cols = []
        if header_idx + 1 < len(raw):
            data_row = raw[header_idx + 1]
            for i, c in enumerate(data_row):
                if i > 0:
                    try:
                        float(str(c if c is not None else "").replace(",", "").replace(" ", ""))
                        numeric_cols.append(i)
                    except (ValueError, TypeError):
                        pass
        if len(numeric_cols) >= 4:
            col_prev, col_deposit, col_withdrawal, col_balance = numeric_cols[:4]
        elif len(numeric_cols) >= 2:
            col_deposit = numeric_cols[0]
            col_withdrawal = numeric_cols[1]
            if len(numeric_cols) >= 3:
                col_balance = numeric_cols[2]

    if col_account < 0:
        col_account = 0

    print(f"[Fund] columns: account={col_account}, deposit={col_deposit}, "
          f"withdrawal={col_withdrawal}, balance={col_balance}, prev={col_prev}")

    # ── 날짜 추출 ──
    for i in range(min(header_idx, 5)):
        row_text = " ".join(str(c if c is not None else "") for c in raw[i])
        m = re.search(r"(\d{4})[.\-/년]\s*(\d{1,2})[.\-/월]\s*(\d{1,2})", row_text)
        if m:
            result["date"] = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            break

    # ── 데이터 행 파싱 ──
    summary_pat = re.compile(r"합\s*계|총\s*계|소\s*계|total|subtotal", re.IGNORECASE)
    for i in range(header_idx + 1, len(raw)):
        row = raw[i]
        row_text = " ".join(str(c if c is not None else "") for c in row)
        if summary_pat.search(row_text):
            continue
        if not any(c is not None and str(c).strip() for c in row):
            continue

        deposit = parse_amount(row[col_deposit]) if col_deposit >= 0 and col_deposit < len(row) else 0
        withdrawal = parse_amount(row[col_withdrawal]) if col_withdrawal >= 0 and col_withdrawal < len(row) else 0
        prev_bal = parse_amount(row[col_prev]) if col_prev >= 0 and col_prev < len(row) else 0
        balance = parse_amount(row[col_balance]) if col_balance >= 0 and col_balance < len(row) else 0

        if deposit == 0 and withdrawal == 0 and balance == 0 and prev_bal == 0:
            continue

        acct = str(row[col_account] if col_account >= 0 and col_account < len(row) and row[col_account] is not None else "").strip()
        bank = str(row[col_bank] if col_bank >= 0 and col_bank < len(row) and row[col_bank] is not None else "").strip()
        label = f"{bank} {acct}".strip() if bank else (acct or f"항목{len(result['transactions'])+1}")

        result["transactions"].append({
            "account": label, "prevBalance": prev_bal, "deposit": deposit,
            "withdrawal": withdrawal, "balance": balance, "description": label,
        })

    return result


def auto_detect_bank_mapping(headers):
    date_keys = ["거래일", "거래일자", "일자", "날짜", "date", "거래일시", "처리일"]
    deposit_keys = ["입금", "입금액", "입금금액", "입금(원)", "대변", "credit", "받은금액"]
    withdrawal_keys = ["출금", "출금액", "출금금액", "출금(원)", "차변", "debit", "보낸금액", "지급액"]
    balance_keys = ["잔액", "잔고", "거래후잔액", "balance", "현재잔액", "통장잔액"]
    desc_keys = ["적요", "거래내용", "내용", "비고", "메모", "description", "거래구분", "처리내용"]
    acct_keys = ["계좌", "계좌번호", "은행", "은행명", "account"]

    headers_lower = [h.lower() for h in headers]
    def find(keys):
        for i, h in enumerate(headers_lower):
            if any(k.lower() in h for k in keys):
                return i
        return -1

    return {
        "date": find(date_keys),
        "deposit": find(deposit_keys),
        "withdrawal": find(withdrawal_keys),
        "balance": find(balance_keys),
        "description": find(desc_keys),
        "account": find(acct_keys),
    }


def parse_bank_excel(raw):
    """handleBankFile의 Excel 분기 + parseBankRows 재현 (합계 행 필터 포함)."""
    if not raw or len(raw) < 2:
        return {"transactions": [], "headers": []}

    header_idx = 0
    for i in range(min(len(raw), 10)):
        non_empty = sum(1 for c in raw[i] if c is not None and str(c).strip())
        if non_empty >= 3:
            header_idx = i
            break

    headers = [str(c if c is not None else "").strip() for c in raw[header_idx]]
    print(f"[Bank] headerIdx={header_idx}, headers={headers}")
    rows = [r for r in raw[header_idx + 1:] if any(c is not None and str(c).strip() for c in r)]

    mapping = auto_detect_bank_mapping(headers)
    print(f"[Bank] mapping={mapping}")

    summary_pat = re.compile(r"합\s*계|총\s*계|소\s*계|^\s*total\s*$|^\s*subtotal\s*$|이월|이전잔액", re.IGNORECASE)
    transactions = []
    for row in rows:
        row_text = " ".join(str(c if c is not None else "") for c in row)
        if summary_pat.search(row_text):
            continue
        deposit = parse_amount(row[mapping["deposit"]]) if mapping["deposit"] >= 0 else 0
        withdrawal = parse_amount(row[mapping["withdrawal"]]) if mapping["withdrawal"] >= 0 else 0
        if deposit == 0 and withdrawal == 0:
            continue
        transactions.append({
            "date": str(row[mapping["date"]] if mapping["date"] >= 0 else ""),
            "deposit": deposit,
            "withdrawal": withdrawal,
            "balance": parse_amount(row[mapping["balance"]]) if mapping["balance"] >= 0 else None,
            "description": str(row[mapping["description"]] if mapping["description"] >= 0 else ""),
        })

    return {"transactions": transactions, "headers": headers}


def find_subset_sum(amounts, target, tolerance):
    """findSubsetSum greedy 알고리즘 그대로 재현."""
    if not amounts:
        return None
    s = sorted(amounts, reverse=True)
    remaining = target
    selected = []
    for amt in s:
        if amt <= remaining + tolerance:
            selected.append(amt)
            remaining -= amt
            if abs(remaining) <= tolerance:
                return selected
    return selected if abs(remaining) <= tolerance else None


def cross_validate(fund_data, bank_data, tolerance=10):
    fund_tx = fund_data["transactions"]
    bank_tx = bank_data["transactions"]

    fund_dep = sum(t["deposit"] for t in fund_tx)
    fund_wd = sum(t["withdrawal"] for t in fund_tx)
    bank_dep = sum(t["deposit"] for t in bank_tx)
    bank_wd = sum(t["withdrawal"] for t in bank_tx)

    print()
    print("=" * 60)
    print("[Step 2: 합계 검증]")
    ok_dep = abs(fund_dep - bank_dep) <= tolerance
    ok_wd = abs(fund_wd - bank_wd) <= tolerance
    print(f"  자금일보 수입: {fund_dep:>15,.0f}원")
    print(f"  은행     입금: {bank_dep:>15,.0f}원   {'[OK]' if ok_dep else '[X] 차이 ' + format(fund_dep-bank_dep, ',')}")
    print(f"  자금일보 지출: {fund_wd:>15,.0f}원")
    print(f"  은행     출금: {bank_wd:>15,.0f}원   {'[OK]' if ok_wd else '[X] 차이 ' + format(fund_wd-bank_wd, ',')}")

    bank_matched = set()
    fund_matched = set()
    inline_results = []

    # === Step 4: 1:1 정확 매칭 ===
    print()
    print("[Step 4: 1:1 정확 매칭]")
    for fi, ft in enumerate(fund_tx):
        fd = ft["deposit"]
        fw = ft["withdrawal"]
        if fd == 0 and fw == 0:
            fund_matched.add(fi)
            continue
        best_match, best_diff, best_is_dep = -1, float("inf"), True
        for bi, bt in enumerate(bank_tx):
            if bi in bank_matched:
                continue
            if fd > 0 and bt["deposit"] > 0:
                diff = abs(fd - bt["deposit"])
                if diff <= tolerance and diff < best_diff:
                    best_diff, best_match, best_is_dep = diff, bi, True
            if fw > 0 and bt["withdrawal"] > 0:
                diff = abs(fw - bt["withdrawal"])
                if diff <= tolerance and diff < best_diff:
                    best_diff, best_match, best_is_dep = diff, bi, False
        if best_match >= 0:
            bank_matched.add(best_match)
            fund_matched.add(fi)
            bt = bank_tx[best_match]
            kind = "수입" if best_is_dep else "지출"
            amt = fd if best_is_dep else fw
            print(f"  [OK] {ft['account']:25} {kind} {amt:>12,.0f}원  <->  은행 {bt['description']}")
            inline_results.append({"status": "match", "fund": ft["account"], "bank": bt["description"]})

    # === Step 5: N:1 합산 매칭 (subset sum) ===
    print()
    print("[Step 5: N:1 합산 매칭 (서브셋 합)]")
    for fi, ft in enumerate(fund_tx):
        if fi in fund_matched:
            continue
        fd, fw = ft["deposit"], ft["withdrawal"]
        if fd == 0 and fw == 0:
            continue
        target = fd if fd > 0 else fw
        is_dep = fd > 0
        candidates = []
        for bi, bt in enumerate(bank_tx):
            if bi in bank_matched:
                continue
            ba = bt["deposit"] if is_dep else bt["withdrawal"]
            if ba > 0:
                candidates.append({"idx": bi, "amount": ba, "desc": bt["description"]})

        cand_sum = sum(c["amount"] for c in candidates)
        if abs(cand_sum - target) <= tolerance and candidates:
            for c in candidates:
                bank_matched.add(c["idx"])
            fund_matched.add(fi)
            descs = ", ".join(f"{c['desc']}({c['amount']:,.0f})" for c in candidates)
            print(f"  [OK] {ft['account']:25} {target:>12,.0f}원  <->  은행 {len(candidates)}건 합산 [{descs}]")
            inline_results.append({"status": "group", "fund": ft["account"], "n": len(candidates)})
        elif candidates and len(candidates) <= 30:
            found = find_subset_sum([c["amount"] for c in candidates], target, tolerance)
            if found:
                used = []
                local_matched = set()
                for amt in found:
                    for c in candidates:
                        if c["amount"] == amt and c["idx"] not in bank_matched and c["idx"] not in local_matched:
                            local_matched.add(c["idx"])
                            used.append(c)
                            break
                if used:
                    for c in used:
                        bank_matched.add(c["idx"])
                    fund_matched.add(fi)
                    descs = ", ".join(f"{c['desc']}({c['amount']:,.0f})" for c in used)
                    print(f"  [OK] {ft['account']:25} {target:>12,.0f}원  <->  은행 {len(used)}건 서브셋합 [{descs}]")
                    inline_results.append({"status": "subset", "fund": ft["account"], "n": len(used)})
                else:
                    print(f"  [X] {ft['account']:25} {target:>12,.0f}원  서브셋 합산 매칭 실패")
            else:
                print(f"  [X] {ft['account']:25} {target:>12,.0f}원  서브셋 합 못 찾음")

    # 미매칭 결과
    print()
    print("[미매칭 자금일보 항목]")
    any_unmatched_fund = False
    for fi, ft in enumerate(fund_tx):
        if fi in fund_matched:
            continue
        if ft["deposit"] == 0 and ft["withdrawal"] == 0:
            continue
        any_unmatched_fund = True
        amt = ft["deposit"] or ft["withdrawal"]
        print(f"  [X] {ft['account']} {amt:,.0f}원")
    if not any_unmatched_fund:
        print("  (없음 - 모든 자금일보 항목이 매칭됨)")

    print()
    print("[미매칭 은행 거래]")
    any_unmatched_bank = False
    for bi, bt in enumerate(bank_tx):
        if bi in bank_matched:
            continue
        if bt["deposit"] == 0 and bt["withdrawal"] == 0:
            continue
        any_unmatched_bank = True
        amt = bt["deposit"] or bt["withdrawal"]
        print(f"  [X] {bt['description']} {amt:,.0f}원")
    if not any_unmatched_bank:
        print("  (없음 - 모든 은행 거래가 매칭됨)")

    print()
    print("=" * 60)
    if ok_dep and ok_wd and not any_unmatched_fund and not any_unmatched_bank:
        print("최종 결과: [성공] 모든 검증 통과 - 일치율 100%")
    else:
        print("최종 결과: [실패] 검증 오류 있음")


def load_xlsx(path):
    wb = load_workbook(path)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


if __name__ == "__main__":
    print("=" * 60)
    print("자금일보 파싱 시뮬레이션")
    print("=" * 60)
    fund_raw = load_xlsx(OUT / "자금일보_샘플.xlsx")
    fund_data = parse_fund_excel(fund_raw)
    print(f"[Fund] date={fund_data['date']}")
    print(f"[Fund] {len(fund_data['transactions'])}건의 거래:")
    for t in fund_data["transactions"]:
        print(f"  - {t['account']:30} 수입={t['deposit']:>13,.0f}  지출={t['withdrawal']:>13,.0f}")

    print()
    print("=" * 60)
    print("은행 거래내역 파싱 시뮬레이션")
    print("=" * 60)
    bank_raw = load_xlsx(OUT / "은행거래내역_샘플.xlsx")
    bank_data = parse_bank_excel(bank_raw)
    print(f"[Bank] {len(bank_data['transactions'])}건의 거래:")
    for t in bank_data["transactions"]:
        print(f"  - {t['date']:20} 입금={t['deposit']:>13,.0f}  출금={t['withdrawal']:>13,.0f}  {t['description']}")

    cross_validate(fund_data, bank_data)
