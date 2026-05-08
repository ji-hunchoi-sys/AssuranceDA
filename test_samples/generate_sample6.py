# -*- coding: utf-8 -*-
"""
자금일보 검증기 샘플6 생성 — 4개 은행 분리 + 통합 자금일보.

시나리오 (2026-05-07, (주)한국상사):
  · 신한은행 주계좌 — 매출 회수 + 거래처 결제 + 급여 지급 + 소득세 (5건)
  · 국민은행 운영계좌 — 외상대 결제 + 부가세 (3건)
  · 하나은행 보조계좌 — 임대료/통신비/사무용품 (3건)
  · 우리은행 외화계좌 — 외화 매출 + 환차손 (2건)
  → 총 13건 거래 (입금 3 + 출금 10)

생성 파일 (총 10개):
  자금일보:
    · 자금일보_샘플6.xlsx, .pdf

  은행거래내역 (4개 은행 × 2포맷):
    · 은행거래내역_샘플6_신한.xlsx, .pdf
    · 은행거래내역_샘플6_국민.xlsx, .pdf
    · 은행거래내역_샘플6_하나.xlsx, .pdf
    · 은행거래내역_샘플6_우리.xlsx, .pdf

검증 시나리오:
  1) 자금일보 검증기에 자금일보 1개 + 은행 4개 누적 업로드 → 13건 1:1 매칭
  2) 자금일보 자동 생성기에 은행 4개 입력 → 1개 자금일보 (13건 거래)
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

OUT_DIR = Path(__file__).resolve().parent

THIN = Side(style="thin", color="888888")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="E8EEF5")
SECTION_FILL = PatternFill("solid", fgColor="F4F6F9")
INCOME_FILL = PatternFill("solid", fgColor="ECFDF5")
EXPENSE_FILL = PatternFill("solid", fgColor="FEF2F2")


def register_korean_font():
    candidates = [
        ("MalgunGothic", r"C:\Windows\Fonts\malgun.ttf"),
        ("MalgunGothicBold", r"C:\Windows\Fonts\malgunbd.ttf"),
        ("NanumGothic", r"C:\Windows\Fonts\NanumGothic.ttf"),
    ]
    registered = []
    for name, path in candidates:
        if Path(path).exists():
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                registered.append(name)
            except Exception as e:
                print(f"  [WARN] font register failed: {name}: {e}")
    if not registered:
        raise RuntimeError("한글 폰트를 찾지 못했습니다 (Malgun/Nanum)")
    return registered[0], (registered[1] if len(registered) > 1 else registered[0])


KOR_FONT, KOR_FONT_BOLD = register_korean_font()


def styled(font=None, size=10, bold=False, align="LEFT"):
    return ParagraphStyle(
        "k",
        fontName=KOR_FONT_BOLD if bold else (font or KOR_FONT),
        fontSize=size,
        leading=size * 1.4,
        alignment={"LEFT": 0, "CENTER": 1, "RIGHT": 2}[align],
    )


# ════════════════════════════════════════════════════════════════════
# 시나리오 데이터
# ════════════════════════════════════════════════════════════════════
SCENARIO_DATE = "2026-05-07"
COMPANY = "(주)한국상사"

# (시간, 입금/출금, 계정과목, 거래처, 적요, 입금, 출금)
SHINHAN_TXNS = [
    ("09:30:14", "입금", "매출수금",     "(주)A상사",      "1분기 매출채권 회수", 80_000_000, 0),
    ("10:15:42", "입금", "매출수금",     "(주)B산업",      "4월 매출채권 회수",   35_000_000, 0),
    ("11:30:08", "출금", "외상대 결제",  "(주)C기업",      "원자재 매입 결제",    0, 50_000_000),
    ("14:00:33", "출금", "인건비",       "임직원 일괄",    "5월 급여 지급",       0, 120_000_000),
    ("16:00:11", "출금", "세금",         "강남세무서",     "4월 소득세 납부",     0, 8_500_000),
]
KOOKMIN_TXNS = [
    ("10:00:25", "출금", "외상대 결제",  "(주)D상사",      "외상대 결제",         0, 25_000_000),
    ("13:30:50", "출금", "외상대 결제",  "(주)E산업",      "외상대 결제",         0, 15_000_000),
    ("15:00:18", "출금", "세금",         "강남세무서",     "4월 부가세 납부",     0, 12_000_000),
]
HANA_TXNS = [
    ("11:00:22", "출금", "일반경비",     "한빛빌딩",       "임대료 지급",         0, 5_000_000),
    ("11:15:08", "출금", "일반경비",     "KT",             "통신비 자동납부",     0, 800_000),
    ("11:30:55", "출금", "일반경비",     "오피스월드",     "사무용품 구입",       0, 1_200_000),
]
WOORI_TXNS = [
    ("09:45:33", "입금", "외화매출",     "해외바이어",     "외화 매출대금 입금",  20_000_000, 0),
    ("14:30:14", "출금", "환차손",       "환율정산",       "외화 환산차손 지급",  0, 500_000),
]

BANKS = [
    ("신한", "신한은행 주계좌",   "456-789012-01-001", 500_000_000, SHINHAN_TXNS),
    ("국민", "국민은행 운영계좌", "789-012345-02-002", 200_000_000, KOOKMIN_TXNS),
    ("하나", "하나은행 보조계좌", "123-456789-03-003", 100_000_000, HANA_TXNS),
    ("우리", "우리은행 외화계좌", "012-345678-04-004", 50_000_000,  WOORI_TXNS),
]

# 자금일보 — 4개 은행 거래 통합 + 거래은행 컬럼
ALL_FUND_TXNS = []
for (key, bank_label, acct_no, opening, txns) in BANKS:
    for (t, kind, account, vendor, memo, dep, wd) in txns:
        ALL_FUND_TXNS.append((t, kind, account, vendor, bank_label, memo, dep, wd))

# 시간순 정렬
ALL_FUND_TXNS.sort(key=lambda x: x[0])

SUM_DEP = sum(t[6] for t in ALL_FUND_TXNS)
SUM_WD = sum(t[7] for t in ALL_FUND_TXNS)


# ════════════════════════════════════════════════════════════════════
# 1) 자금일보 Excel — 분류형 + 거래은행 컬럼
# ════════════════════════════════════════════════════════════════════
def build_fund_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "자금일보"

    ws["A1"] = "일일 자금일보 (다은행 통합)"
    ws["A1"].font = Font(size=15, bold=True)
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center")

    closing_total = sum(b[3] for b in BANKS) + SUM_DEP - SUM_WD
    opening_total = sum(b[3] for b in BANKS)
    meta = [
        ("회사명", COMPANY, "회계일자", SCENARIO_DATE),
        ("기초 시재", opening_total, "마감 시재", closing_total),
    ]
    for r_off, row in enumerate(meta, start=2):
        for c_off, val in enumerate(row, start=1):
            cell = ws.cell(row=r_off, column=c_off, value=val)
            if c_off in (1, 3):
                cell.font = Font(bold=True)
                cell.fill = SECTION_FILL
            elif isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"

    header_row = 5
    headers = ["구분", "계정과목", "거래처", "거래은행", "적요", "입금", "출금"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center")

    for i, (t, kind, account, vendor, bank, memo, dep, wd) in enumerate(
        ALL_FUND_TXNS, start=header_row + 1
    ):
        vals = [kind, account, vendor, bank, memo, dep, wd]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col in (6, 7) and isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0;-#,##0;\"-\""
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")
                cell.fill = INCOME_FILL if kind == "입금" else EXPENSE_FILL
            else:
                cell.alignment = Alignment(horizontal="left")

    # 합계 행
    tot_row = header_row + 1 + len(ALL_FUND_TXNS)
    ws.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=5)
    ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=tot_row, column=6, value=SUM_DEP).number_format = "#,##0"
    ws.cell(row=tot_row, column=7, value=SUM_WD).number_format = "#,##0"
    for col in range(1, 8):
        c = ws.cell(row=tot_row, column=col)
        c.font = Font(bold=True)
        c.fill = SECTION_FILL
        c.border = BOX

    widths = {"A": 8, "B": 14, "C": 16, "D": 22, "E": 28, "F": 16, "G": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = OUT_DIR / "자금일보_샘플6.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")
    return out


# ════════════════════════════════════════════════════════════════════
# 2) 자금일보 PDF
# ════════════════════════════════════════════════════════════════════
def build_fund_pdf():
    out = OUT_DIR / "자금일보_샘플6.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=18 * mm, bottomMargin=15 * mm,
    )

    title = styled(size=17, bold=True, align="CENTER")
    sub = styled(size=10, align="CENTER")
    note = styled(size=9)

    closing_total = sum(b[3] for b in BANKS) + SUM_DEP - SUM_WD
    opening_total = sum(b[3] for b in BANKS)

    story = [
        Paragraph("일일 자금일보 (다은행 통합)", title),
        Spacer(1, 4 * mm),
        Paragraph(
            f"법인명: {COMPANY}   |   회계일자: {SCENARIO_DATE}   |   "
            f"포함 은행: 신한·국민·하나·우리 (4개)",
            sub,
        ),
        Spacer(1, 6 * mm),
    ]

    meta = Table(
        [["기초 시재", f"{opening_total:,}원",
          "마감 시재", f"{closing_total:,}원"]],
        colWidths=[45 * mm, 45 * mm, 45 * mm, 45 * mm],
    )
    meta.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), KOR_FONT_BOLD, 11),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#E8EEF5")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#E8EEF5")),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("ALIGN", (3, 0), (3, 0), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(meta)
    story.append(Spacer(1, 6 * mm))

    # 은행별 잔액 ※ 메모
    for (key, bank_label, acct_no, opening, txns) in BANKS:
        bank_dep = sum(t[5] for t in txns)
        bank_wd = sum(t[6] for t in txns)
        bank_closing = opening + bank_dep - bank_wd
        story.append(Paragraph(
            f"※ {bank_label} ({acct_no}): 기초 {opening:,}원 → 마감 {bank_closing:,}원 "
            f"(입금 {bank_dep:,} / 출금 {bank_wd:,})",
            note,
        ))
    story.append(Spacer(1, 5 * mm))

    # 입출금내역 통합 표
    fmt = lambda n: f"{n:,}" if isinstance(n, int) and n > 0 else "-"
    headers = ["구분", "계정과목", "거래처", "거래은행", "적요", "입금", "출금"]
    rows_pdf = [headers] + [
        [kind, account, vendor, bank, memo, fmt(dep), fmt(wd)]
        for (t, kind, account, vendor, bank, memo, dep, wd) in ALL_FUND_TXNS
    ]
    # 합계는 표 밖

    table_rows = []
    for i, r in enumerate(rows_pdf):
        cells = []
        for j, c in enumerate(r):
            align = "RIGHT" if j in (5, 6) else ("CENTER" if j == 0 else "LEFT")
            bold = (i == 0)
            cells.append(Paragraph(str(c), styled(size=8, bold=bold, align=align)))
        table_rows.append(cells)

    tbl = Table(
        table_rows,
        colWidths=[14 * mm, 24 * mm, 28 * mm, 30 * mm, 38 * mm, 22 * mm, 22 * mm],
    )
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for i, row in enumerate(ALL_FUND_TXNS, start=1):
        kind = row[1]
        bg = "#ECFDF5" if kind == "입금" else "#FEF2F2"
        style_cmds.append(("BACKGROUND", (0, i), (0, i), colors.HexColor(bg)))
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        f"※ 본 자금일보 — 입금 합계 {SUM_DEP:,}원 / 출금 합계 {SUM_WD:,}원 ({len(ALL_FUND_TXNS)}건)",
        note,
    ))

    doc.build(story)
    print(f"  [OK] {out.name}")
    return out


# ════════════════════════════════════════════════════════════════════
# 3) 은행 거래내역 — 은행별 Excel/PDF (4개 × 2 = 8개 파일)
# ════════════════════════════════════════════════════════════════════
def build_bank_excel(key, bank_label, acct_no, opening, txns):
    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"

    ws["A1"] = f"{bank_label} 거래내역 조회"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"], ws["B2"] = "계좌번호", acct_no
    ws["A3"], ws["B3"] = "예금주", COMPANY
    ws["A4"], ws["B4"] = "조회기간", f"{SCENARIO_DATE} ~ {SCENARIO_DATE}"

    headers = ["거래일시", "입금액", "출금액", "잔액", "적요", "거래처"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center")

    bal = opening
    for i, (t, kind, account, vendor, memo, dep, wd) in enumerate(txns, start=7):
        bal = bal + dep - wd
        date_str = f"{SCENARIO_DATE} {t}"
        vals = [date_str, dep, wd, bal, memo, vendor]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col in (2, 3, 4) and isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

    tot_row = 7 + len(txns)
    ws.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=tot_row, column=2, value=sum(t[5] for t in txns)).number_format = "#,##0"
    ws.cell(row=tot_row, column=3, value=sum(t[6] for t in txns)).number_format = "#,##0"
    for col in range(1, 7):
        c = ws.cell(row=tot_row, column=col)
        c.font = Font(bold=True)
        c.fill = SECTION_FILL
        c.border = BOX

    widths = {"A": 22, "B": 16, "C": 16, "D": 18, "E": 28, "F": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = OUT_DIR / f"은행거래내역_샘플6_{key}.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")


def build_bank_pdf(key, bank_label, acct_no, opening, txns):
    out = OUT_DIR / f"은행거래내역_샘플6_{key}.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    title = styled(size=16, bold=True, align="CENTER")
    sub = styled(size=9, align="CENTER")
    story = [
        Paragraph(f"{bank_label} 거래내역 조회", title),
        Spacer(1, 3 * mm),
        Paragraph(
            f"계좌번호: {acct_no}   |   예금주: {COMPANY}   |   "
            f"조회기간: {SCENARIO_DATE} ~ {SCENARIO_DATE}",
            sub,
        ),
        Spacer(1, 5 * mm),
    ]

    headers = ["거래일시", "입금액", "출금액", "잔액", "적요", "거래처"]
    bal = opening
    data = [headers]
    for (t, kind, account, vendor, memo, dep, wd) in txns:
        bal = bal + dep - wd
        data.append([
            f"{SCENARIO_DATE} {t}",
            f"{dep:,}" if dep else "0",
            f"{wd:,}" if wd else "0",
            f"{bal:,}",
            memo,
            vendor,
        ])
    # 합계는 표 밖

    table_rows = []
    for i, row in enumerate(data):
        cells = []
        for j, c in enumerate(row):
            align = "RIGHT" if j in (1, 2, 3) else ("CENTER" if j == 0 else "LEFT")
            bold = (i == 0)
            cells.append(Paragraph(str(c), styled(size=8, bold=bold, align=align)))
        table_rows.append(cells)

    tbl = Table(
        table_rows,
        colWidths=[34 * mm, 23 * mm, 23 * mm, 28 * mm, 35 * mm, 35 * mm],
    )
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF5")),
        ("BOX", (0, 0), (-1, -1), 0.7, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        f"※ {bank_label} — 입금 {sum(t[5] for t in txns):,}원 / "
        f"출금 {sum(t[6] for t in txns):,}원 ({len(txns)}건)",
        styled(size=9),
    ))

    doc.build(story)
    print(f"  [OK] {out.name}")


# ════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"[샘플6] 회계일자: {SCENARIO_DATE}, 회사: {COMPANY}")
    print(f"  거래 건수: {len(ALL_FUND_TXNS)}건 (입금 {sum(1 for t in ALL_FUND_TXNS if t[1]=='입금')}건 + "
          f"출금 {sum(1 for t in ALL_FUND_TXNS if t[1]=='출금')}건)")
    print(f"  합계: 입금 {SUM_DEP:,}원 / 출금 {SUM_WD:,}원")
    print()

    print("[자금일보] 생성 중...")
    build_fund_excel()
    build_fund_pdf()
    print()

    print("[은행 거래내역] 4개 은행 × 2포맷 생성 중...")
    for (key, bank_label, acct_no, opening, txns) in BANKS:
        build_bank_excel(key, bank_label, acct_no, opening, txns)
        build_bank_pdf(key, bank_label, acct_no, opening, txns)
    print()

    print("=" * 60)
    print("생성 완료. 총 10개 파일")
    print("  · 자금일보_샘플6.xlsx, .pdf")
    print("  · 은행거래내역_샘플6_{신한,국민,하나,우리}.xlsx, .pdf")
