# -*- coding: utf-8 -*-
"""
자금일보 검증기 테스트 파일 생성 스크립트.

생성 파일:
  - 자금일보_샘플.xlsx        (자금일보, 검증 대상)
  - 은행거래내역_샘플.xlsx    (은행 입출금 내역, 검증 기준)

테스트 시나리오 (계획일자 2024-11-11):
  전일이월:  100,000,000원
  수입 합계:  80,000,000원 (외상대 수금 50M + 매출 수금 30M)
  지출 합계:  35,000,000원 (일반경비 5M + 외상대 결제 20M + 전도금 10M)
  조달 후 시재: 145,000,000원

매칭 검증 포인트:
  · 1:1 매칭     : 외상대 수금(50M), 매출 수금(30M), 전도금(10M)
  · N:1 합산매칭 : 일반경비(5M = 1.5+1.5+2M), 외상대 결제(20M = 8+12M)
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pathlib import Path

OUT_DIR = Path(__file__).resolve().parent

THIN = Side(style="thin", color="888888")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="E8EEF5")
SECTION_FILL = PatternFill("solid", fgColor="F4F6F9")


# ─────────────────────────────────────────────────────────────────────
# 1) 자금일보 샘플
# ─────────────────────────────────────────────────────────────────────
def build_fund_report():
    """
    검증기의 parseFundExcel은 전일이월/과부족/조달후 행도 잔액만 있으면
    가짜 거래로 만들어버리므로, 잔액 정보는 메타 영역(상단)에 두고
    데이터 표(헤더 아래)에는 실제 입출금 항목만 둔다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "자금일보"

    # 상단 메타 정보 (parseFundExcel이 데이터로 잘못 읽지 않도록 표 밖)
    ws["A1"] = "일일 자금 계획"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    meta = [
        ("회사명", "(주)테스트회사",   "계획일자", "2024-11-11"),
        ("전일이월 현금시재", 100_000_000, "조달 후 현금시재", 145_000_000),
        ("총 수입", 80_000_000,         "총 지출", 35_000_000),
    ]
    for r_off, row in enumerate(meta, start=2):
        for c_off, val in enumerate(row, start=1):
            cell = ws.cell(row=r_off, column=c_off, value=val)
            if c_off in (1, 3):
                cell.font = Font(bold=True)
                cell.fill = SECTION_FILL
            elif isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"

    # 표 헤더 (6행)
    header_row = 6
    headers = ["구분", "항목", "입금(수입)", "출금(지출)", "비고"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")

    # 실제 입출금 항목만 (잔액 표시 행 제외)
    rows = [
        ("수입", "외상대 수금(원화)", 50_000_000, 0, "고객사A 매출채권 회수"),
        ("수입", "매출 수금",         30_000_000, 0, "고객사B 매출 입금"),
        ("수입", "외상대 수금(외화)",  0,         0, ""),
        ("지출", "차입금 상환",        0,         0, ""),
        ("지출", "외상대 결제(원화)",  0,        20_000_000, "거래처A 8M + 거래처B 12M"),
        ("지출", "외상대 결제(외화)",  0,         0, ""),
        ("지출", "설비투자",           0,         0, ""),
        ("지출", "연구개발비",         0,         0, ""),
        ("지출", "급/상여(회직급여)",  0,         0, ""),
        ("지출", "일반경비",           0,         5_000_000, "사무용품 1.5M + 통신비 1.5M + 임대료 2M"),
        ("지출", "전도금",             0,        10_000_000, "김앤장 법무사사무소"),
    ]

    for i, row in enumerate(rows, start=header_row + 1):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col == 1:
                cell.fill = SECTION_FILL
                cell.alignment = Alignment(horizontal="center")
            elif col == 2:
                cell.alignment = Alignment(horizontal="left")
            elif col in (3, 4):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, int):
                    cell.number_format = "#,##0"
            elif col == 5:
                cell.alignment = Alignment(horizontal="left")
                cell.font = Font(size=10, color="666666")

    # 열 너비
    for col, width in zip("ABCDE", [10, 22, 16, 16, 38]):
        ws.column_dimensions[col].width = width

    out = OUT_DIR / "자금일보_샘플.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")
    return out


# ─────────────────────────────────────────────────────────────────────
# 2) 은행 거래내역 샘플 — 자금일보와 합계가 맞도록 구성
# ─────────────────────────────────────────────────────────────────────
def build_bank_statement():
    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"

    ws["A1"] = "거래내역 조회"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = "계좌번호"
    ws["B2"] = "123-456789-01-001"
    ws["A3"] = "예금주"
    ws["B3"] = "(주)테스트회사"
    ws["A4"] = "조회기간"
    ws["B4"] = "2024-11-11 ~ 2024-11-11"

    headers = ["거래일자", "입금액", "출금액", "잔액", "적요", "거래처", "은행명"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")

    # 잔액 = 100,000,000 (시작)
    # 거래 순서대로 잔액이 변동되도록 구성
    txns = []
    bal = 100_000_000

    def add(deposit, withdrawal, time, desc, vendor):
        nonlocal bal
        bal = bal + deposit - withdrawal
        txns.append((f"2024-11-11 {time}", deposit, withdrawal, bal, desc, vendor, "테스트은행"))

    # 수입 (자금일보: 외상대 수금 50M + 매출 수금 30M)
    add(50_000_000, 0, "09:15:22", "외상대 수금", "고객사A 주식회사")
    add(30_000_000, 0, "10:32:08", "매출 수금",   "고객사B 주식회사")

    # 지출: 일반경비 5M (3건 합산 매칭)
    add(0, 1_500_000, "13:01:11", "사무용품 구입", "오피스디포")
    add(0, 1_500_000, "13:42:50", "통신비 납부",   "KT")
    add(0, 2_000_000, "14:05:33", "임대료 지급",   "한빛빌딩")

    # 지출: 외상대 결제 20M (2건 합산 매칭)
    add(0, 8_000_000,  "15:20:14", "외상대 결제", "거래처A 상사")
    add(0, 12_000_000, "15:21:02", "외상대 결제", "거래처B 상사")

    # 지출: 전도금 10M (1:1 매칭)
    add(0, 10_000_000, "16:55:48", "전도금 지급", "김앤장 법무사사무소")

    for i, t in enumerate(txns, start=7):
        for col, val in enumerate(t, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BOX
            if col in (2, 3, 4) and isinstance(val, int):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
            elif col == 1:
                cell.alignment = Alignment(horizontal="center")

    # 합계 행
    total_row = 7 + len(txns)
    ws.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=sum(t[1] for t in txns)).number_format = "#,##0"
    ws.cell(row=total_row, column=3, value=sum(t[2] for t in txns)).number_format = "#,##0"
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).fill = SECTION_FILL
        ws.cell(row=total_row, column=col).border = BOX

    # 열 너비
    widths = {"A": 20, "B": 14, "C": 14, "D": 16, "E": 18, "F": 22, "G": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = OUT_DIR / "은행거래내역_샘플.xlsx"
    wb.save(out)
    print(f"  [OK] {out.name}")
    return out


if __name__ == "__main__":
    print("테스트 파일 생성 중...")
    build_fund_report()
    build_bank_statement()
    print("\n생성된 파일은 test_samples 폴더에 저장됩니다.")
    print("자금일보_샘플.xlsx 와 은행거래내역_샘플.xlsx 를 검증기에 업로드해 보세요.")
